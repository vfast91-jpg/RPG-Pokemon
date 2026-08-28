#!/usr/bin/env python3
"""Build and validate the Generation-3 Pokémon runtime extension.

This tool compiles the approved vertical Gen-3 spreadsheet into the same
append-only JSON schema already used by the project. It only handles Pokémon
data. Move IDs in learnsets/TM lists are preserved as identifiers; move
definitions and mechanics are deliberately outside this tool's scope.

Usage:
    python tools/build_gen3_registry.py /path/to/Pokemon_Datenbank_Gen3_....xlsx
    python tools/build_gen3_registry.py /path/to/file.xlsx --output-root .
    python tools/build_gen3_registry.py /path/to/file.xlsx --output-root . --repo-root .
"""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment-specific
    raise SystemExit(
        "openpyxl is required for the one-time XLSX import: pip install openpyxl"
    ) from exc

SOURCE_DATE = "2026-08-28"
SOURCE_NAME = "Pokemon_Datenbank_Gen3_BALANCIERT_MIT_LERNLISTEN_2026-08-28.xlsx"
SHEET_NAME = "Pokémon-Datenbank"
EXPECTED_SPECIES = 146
EXPECTED_FAMILIES = 73
EXPECTED_BASE_SPECIES = 282
EXPECTED_BASE_FAMILIES = 129
RUNTIME_SPECIES = EXPECTED_BASE_SPECIES + EXPECTED_SPECIES
RUNTIME_FAMILIES = EXPECTED_BASE_FAMILIES + EXPECTED_FAMILIES
MOVE_ID_RE = re.compile(r"^[a-z0-9_]+$")

REQUIRED_FIELDS = [
    "Schema-Version",
    "Spezies-ID",
    "Pokédex-Nummer",
    "Anzeigename",
    "Pokédex-Kategorie",
    "Asset-ID",
    "Größe (m)",
    "Gewicht (kg)",
    "Typ 1",
    "Typ 2",
    "Basis-KP",
    "Basis-Angriff",
    "Basis-Verteidigung",
    "Basis-Statuswert",
    "Basis-Geschwindigkeit",
    "RPG-Basiswertsumme",
    "Original-Basiswertsumme",
    "Vergleichsbudget 5/6",
    "EP-Kurve",
    "Basis-EP-Ausbeute",
    "Stat-Formel-ID",
    "Fangrate",
    "Speziesfamilien-ID",
    "Familien-Fangrate",
    "Entwickelt sich zu",
    "Entwicklungslevel",
    "Entwicklung verpflichtend?",
    "Entwicklungsmethode",
    "Entwicklungsvoraussetzung / Item",
    "Level-Up-Lernliste",
    "TM-/VM-Lernliste",
    "Sonstige Lernwege",
    "Tags",
    "Notizen",
]


def nullable_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "<null>"}:
        return None
    return text


def number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        value_float = float(value)
        return int(value_float) if value_float.is_integer() else value_float
    value_float = float(str(value).strip().replace(",", "."))
    return int(value_float) if value_float.is_integer() else value_float


def split_moves(text: Any) -> list[str]:
    return [part.strip() for part in str(text or "").split(",") if part.strip()]


def dedupe(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value not in result:
            result.append(value)
    return result


def parse_level_learnset(text: Any, species_id: str) -> dict[str, Any]:
    level_up: OrderedDict[str, list[str]] = OrderedDict()
    evolution_moves: list[str] = []
    relearn: list[str] = []
    state = "level"
    lines = [
        line.strip()
        for line in str(text or "").replace("\r", "").splitlines()
        if line.strip()
    ]

    for line in lines:
        match = re.match(r"^(\d+)\s*:\s*(.+)$", line, re.I)
        if match:
            level = str(int(match.group(1)))
            level_up.setdefault(level, []).extend(split_moves(match.group(2)))
            state = "level"
            continue

        match = re.match(r"^Lv\.?\s*(\d+)\s*:\s*(.+)$", line, re.I)
        if match:
            level = str(int(match.group(1)))
            level_up.setdefault(level, []).extend(split_moves(match.group(2)))
            state = "level"
            continue

        match = re.match(r"^RELEARN/LV1\s*:\s*(.+)$", line, re.I)
        if match:
            relearn.extend(split_moves(match.group(1)))
            state = "relearn"
            continue

        match = re.match(r"^EVOLUTION\s*:\s*(.+)$", line, re.I)
        if match:
            evolution_moves.extend(split_moves(match.group(1)))
            state = "evolution"
            continue

        match = re.match(r"^Entwicklungsattacke\s*:\s*(.+)$", line, re.I)
        if match:
            evolution_moves.extend(split_moves(match.group(1)))
            state = "evolution"
            continue

        lower = line.lower()
        if lower in {"level-up", "level-up – bdsp", "level-up - bdsp", "danach"}:
            state = "level"
            continue
        if lower == "entwicklungsattacken":
            state = "evolution"
            continue
        if "relearn" in lower or lower == "bei entwicklung verfügbar":
            state = "relearn"
            continue
        if lower.startswith("keine weitere reguläre levelprogression"):
            continue

        if (
            state in {"relearn", "evolution"}
            and re.fullmatch(r"[a-z0-9_]+(?:\s*,\s*[a-z0-9_]+)*", line)
        ):
            destination = relearn if state == "relearn" else evolution_moves
            destination.extend(split_moves(line))
            continue

        raise ValueError(f"{species_id}: unknown learnset line: {line!r}")

    for level in list(level_up):
        level_up[level] = dedupe(level_up[level])

    return {
        "level_up": level_up,
        "evolution_moves": dedupe(evolution_moves),
        "relearn_lv1": dedupe(relearn),
    }


def parse_tm(text: Any) -> list[str]:
    return [
        line.strip()
        for line in str(text or "").replace("\r", "").splitlines()
        if line.strip()
    ]


def parse_tags(text: Any) -> list[str]:
    return [
        part.strip()
        for part in re.split(r"[;\n]+", str(text or ""))
        if part.strip()
    ]


def evolution_record(row: dict[str, Any]) -> dict[str, Any]:
    target = nullable_text(row["Entwickelt sich zu"])
    level_value = number(row["Entwicklungslevel"])
    level = int(level_value) if level_value is not None else None
    mandatory = bool(row["Entwicklung verpflichtend?"])
    method = nullable_text(row["Entwicklungsmethode"]) or "none"
    requirement = nullable_text(row["Entwicklungsvoraussetzung / Item"])

    if target is None:
        return {
            "evolves_into": None,
            "evolution_level": None,
            "mandatory": False,
            "method": "none",
            "requirement": None,
        }

    if "/" in target:
        targets = [part.strip() for part in target.split("/") if part.strip()]
        return {
            "evolves_into": targets,
            "evolution_level": level,
            "mandatory": mandatory,
            "method": method,
            "requirement": requirement,
            "choices": [
                {"target": choice, "level": level, "method": method}
                for choice in targets
            ],
        }

    return {
        "evolves_into": target,
        "evolution_level": level,
        "mandatory": mandatory,
        "method": method,
        "requirement": requirement,
    }


def load_vertical_sheet(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"sheet {SHEET_NAME!r} is missing")
    sheet = workbook[SHEET_NAME]

    rows = list(sheet.iter_rows(values_only=True))
    labels = [row[0] for row in rows]
    index = {str(label): row_index for row_index, label in enumerate(labels) if label}

    missing = [field for field in REQUIRED_FIELDS if field not in index]
    if missing:
        raise ValueError(f"missing spreadsheet fields: {', '.join(missing)}")

    species_rows: list[dict[str, Any]] = []
    max_column = max(len(row) for row in rows)
    for column in range(1, max_column):
        species_id = rows[index["Spezies-ID"]][column]
        if not species_id:
            continue
        species_rows.append(
            {
                field: rows[index[field]][column]
                for field in REQUIRED_FIELDS
            }
        )
    return species_rows


def build_records(rows: list[dict[str, Any]]) -> tuple[dict[str, Any], ...]:
    core: OrderedDict[str, Any] = OrderedDict()
    details: OrderedDict[str, Any] = OrderedDict()
    families: OrderedDict[str, list[str]] = OrderedDict()

    for row in rows:
        species_id = str(row["Spezies-ID"]).strip()
        if species_id in core:
            raise ValueError(f"duplicate species id: {species_id}")
        asset_id = str(row["Asset-ID"]).strip()
        if not asset_id:
            raise ValueError(f"{species_id}: missing asset id")

        learnset = parse_level_learnset(row["Level-Up-Lernliste"], species_id)
        learnset["tm_hm"] = parse_tm(row["TM-/VM-Lernliste"])

        base_stats = {
            "hp": int(number(row["Basis-KP"])),
            "attack": int(number(row["Basis-Angriff"])),
            "defense": int(number(row["Basis-Verteidigung"])),
            "special": int(number(row["Basis-Statuswert"])),
            "speed": int(number(row["Basis-Geschwindigkeit"])),
        }
        types = {
            "primary": str(row["Typ 1"]).strip(),
            "secondary": nullable_text(row["Typ 2"]),
        }

        core[species_id] = {
            "species_id": species_id,
            "pokedex_number": int(number(row["Pokédex-Nummer"])),
            "display_name": str(row["Anzeigename"]).strip(),
            "asset_id": asset_id,
            "types": types,
            "base_stats": base_stats,
        }

        details[species_id] = {
            "schema_version": str(row["Schema-Version"]),
            "species_id": species_id,
            "pokedex_number": int(number(row["Pokédex-Nummer"])),
            "dex_number": int(number(row["Pokédex-Nummer"])),
            "display_name": str(row["Anzeigename"]).strip(),
            "category": str(row["Pokédex-Kategorie"] or "PENDING"),
            "asset_id": asset_id,
            "height_m": number(row["Größe (m)"]),
            "weight_kg": number(row["Gewicht (kg)"]),
            "types": types,
            "base_stats": base_stats,
            "rpg_basis_sum": number(row["RPG-Basiswertsumme"]),
            "original_bst": number(row["Original-Basiswertsumme"]),
            "comparison_budget": number(row["Vergleichsbudget 5/6"]),
            "experience_curve": str(row["EP-Kurve"]),
            "base_experience": int(number(row["Basis-EP-Ausbeute"])),
            "stat_formula_id": str(row["Stat-Formel-ID"]),
            "catch_rate": number(row["Fangrate"]),
            "family_id": str(row["Speziesfamilien-ID"]),
            "family_catch_rate": number(row["Familien-Fangrate"]),
            "evolution": evolution_record(row),
            "learnset": learnset,
            "tags": parse_tags(row["Tags"]),
            "notes": str(row["Notizen"] or ""),
            "other_learning_paths": str(row["Sonstige Lernwege"] or "none"),
        }

        family_id = str(row["Speziesfamilien-ID"]).strip()
        families.setdefault(family_id, []).append(species_id)

    route_roots = list(families)
    master_pack = {
        "schema_version": 1,
        "source_date": SOURCE_DATE,
        "source": SOURCE_NAME,
        "scope": "Generation 3 append-only core roster extension; Pokémon and forms only.",
        "species_count": len(core),
        "species": core,
    }
    family_pack = {
        "schema_version": 1,
        "source_date": SOURCE_DATE,
        "source": SOURCE_NAME,
        "family_count": len(families),
        "species_count": len(core),
        "route_roots": route_roots,
        "family_members": families,
    }
    detail_packs: list[dict[str, Any]] = []
    detail_paths: list[str] = []
    for start in range(0, len(route_roots), 4):
        shard_roots = route_roots[start:start + 4]
        shard_species: OrderedDict[str, Any] = OrderedDict(
            (species_id, entry)
            for species_id, entry in details.items()
            if entry["family_id"] in shard_roots
        )
        first = start + 1
        last = start + len(shard_roots)
        filename = f"gen3_species_families_{first:02d}_{last:02d}_v1.json"
        detail_paths.append("res://data/" + filename)
        detail_packs.append({
            "schema_version": 12,
            "source_date": SOURCE_DATE,
            "source": SOURCE_NAME,
            "scope": (
                "Generation 3 Pokémon/form detail shard; attack IDs are preserved "
                "as data only and do not implement move mechanics."
            ),
            "family_roots": shard_roots,
            "species_count": len(shard_species),
            "species": shard_species,
        })

    extension_pack = {
        "schema_version": 1,
        "source_date": SOURCE_DATE,
        "scope": "Generation 3 append-only runtime roster extension; Pokémon and forms only.",
        "expected_base_species_count": EXPECTED_BASE_SPECIES,
        "expected_base_route_root_count": EXPECTED_BASE_FAMILIES,
        "extension_species_count": len(core),
        "extension_route_root_count": len(families),
        "runtime_species_count": EXPECTED_BASE_SPECIES + len(core),
        "runtime_route_root_count": EXPECTED_BASE_FAMILIES + len(families),
        "species_master_extension_file": (
            "res://data/pokemon_species_master_extension_gen3_v1.json"
        ),
        "family_meta_extension_file": (
            "res://data/pokemon_family_meta_extension_gen3_v1.json"
        ),
        "species_detail_files": detail_paths,
    }
    return master_pack, family_pack, detail_packs, extension_pack


def validate(
    master_pack: dict[str, Any],
    family_pack: dict[str, Any],
    detail_packs: list[dict[str, Any]],
    extension_pack: dict[str, Any],
    repo_root: Path | None = None,
) -> None:
    core = master_pack["species"]
    details: OrderedDict[str, Any] = OrderedDict()
    for detail_pack in detail_packs:
        for species_id, entry in detail_pack["species"].items():
            if species_id in details:
                raise ValueError(f"duplicate detail species id: {species_id}")
            details[species_id] = entry
    families = family_pack["family_members"]
    roots = family_pack["route_roots"]

    if len(core) != EXPECTED_SPECIES:
        raise ValueError(f"expected {EXPECTED_SPECIES} species/forms, got {len(core)}")
    if len(families) != EXPECTED_FAMILIES:
        raise ValueError(f"expected {EXPECTED_FAMILIES} families, got {len(families)}")
    if set(core) != set(details):
        raise ValueError("core/detail species coverage differs")
    if len(roots) != len(set(roots)):
        raise ValueError("duplicate family root")
    if set(roots) != set(families):
        raise ValueError("route roots and family map keys differ")

    family_species: list[str] = []
    for root in roots:
        members = families[root]
        if root not in members:
            raise ValueError(f"{root}: family root missing from own member list")
        family_species.extend(members)
    if len(family_species) != len(set(family_species)):
        raise ValueError("a species is assigned to more than one family")
    if set(family_species) != set(core):
        raise ValueError("family graph does not cover exactly the Gen-3 extension")

    for species_id, entry in details.items():
        if entry["species_id"] != species_id:
            raise ValueError(f"{species_id}: mismatched species_id")
        if entry["family_id"] not in families:
            raise ValueError(f"{species_id}: unknown family {entry['family_id']}")
        if any(float(entry["base_stats"][key]) <= 0 for key in (
            "hp", "attack", "defense", "special", "speed"
        )):
            raise ValueError(f"{species_id}: invalid base stat")

        evolution = entry["evolution"]
        target = evolution["evolves_into"]
        targets = target if isinstance(target, list) else ([] if target is None else [target])
        for target_id in targets:
            if target_id not in core:
                raise ValueError(f"{species_id}: evolution target missing: {target_id}")

        learnset = entry["learnset"]
        move_groups: list[str] = []
        for moves in learnset["level_up"].values():
            move_groups.extend(moves)
        move_groups.extend(learnset["evolution_moves"])
        move_groups.extend(learnset["relearn_lv1"])
        move_groups.extend(learnset["tm_hm"])
        for move_id in move_groups:
            if not MOVE_ID_RE.fullmatch(move_id):
                raise ValueError(f"{species_id}: invalid move identifier {move_id!r}")

        if repo_root is not None:
            name = entry["display_name"]
            candidates = [
                repo_root / "assets" / f"{name}.{ext}"
                for ext in ("png", "webp", "jpg", "jpeg", "svg")
            ] + [
                repo_root / "assets" / "monsters" / f"{name}.{ext}"
                for ext in ("png", "webp", "jpg", "jpeg", "svg")
            ]
            if not any(path.exists() for path in candidates):
                raise ValueError(f"{species_id}: sprite asset missing for {name!r}")

    if extension_pack["extension_species_count"] != EXPECTED_SPECIES:
        raise ValueError("extension species count mismatch")
    if extension_pack["extension_route_root_count"] != EXPECTED_FAMILIES:
        raise ValueError("extension family count mismatch")
    if extension_pack["runtime_species_count"] != RUNTIME_SPECIES:
        raise ValueError("runtime species count mismatch")
    if extension_pack["runtime_route_root_count"] != RUNTIME_FAMILIES:
        raise ValueError("runtime family count mismatch")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--output-root", type=Path, default=Path("."))
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=None,
        help="Optional repository root for sprite existence validation.",
    )
    args = parser.parse_args()

    rows = load_vertical_sheet(args.xlsx)
    packs = build_records(rows)
    validate(*packs, repo_root=args.repo_root)

    data_dir = args.output_root / "data"
    master_pack, family_pack, detail_packs, extension_pack = packs
    write_json(data_dir / "pokemon_species_master_extension_gen3_v1.json", master_pack)
    write_json(data_dir / "pokemon_family_meta_extension_gen3_v1.json", family_pack)
    for path, payload in zip(extension_pack["species_detail_files"], detail_packs):
        write_json(data_dir / Path(path).name, payload)
    write_json(data_dir / "pokemon_database_extension_gen3_v1.json", extension_pack)

    print(
        f"Gen3 import OK: {EXPECTED_SPECIES} species/forms, "
        f"{EXPECTED_FAMILIES} families, runtime target "
        f"{RUNTIME_SPECIES}/{RUNTIME_FAMILIES}."
    )


if __name__ == "__main__":
    main()
